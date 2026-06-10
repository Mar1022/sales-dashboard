import os
import sys
import json
import csv
import io
import traceback
from datetime import datetime, timedelta
from flask import Flask, render_template, request, jsonify
import psycopg2
import psycopg2.extras
from psycopg2.extras import execute_values
import openpyxl
from urllib.parse import urlparse

app = Flask(__name__)

# ── 数据库配置 ──
DATABASE_URL = os.environ.get('DATABASE_URL', '')
UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)



def get_db():
    """获取 PostgreSQL 连接"""
    conn = psycopg2.connect(DATABASE_URL)
    return conn


def get_cursor(conn):
    """获取字典游标"""
    return conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)


# ── 错误日志 ──
@app.errorhandler(Exception)
def handle_error(e):
    print("=" * 50, flush=True)
    print(f"ERROR: {e}", flush=True)
    traceback.print_exc(file=sys.stdout)
    print("=" * 50, flush=True)
    return jsonify({"error": str(e)}), 500

@app.errorhandler(500)
def handle_500(e):
    print("=" * 50, flush=True)
    print("500 INTERNAL SERVER ERROR", flush=True)
    if hasattr(e, '__traceback__'):
        traceback.print_tb(e.__traceback__, file=sys.stdout)
    else:
        traceback.print_exc(file=sys.stdout)
    print("=" * 50, flush=True)
    return jsonify({"error": "服务器内部错误"}), 500


# ── 默认新品规格清单 ──
DEFAULT_NEW_PRODS = [
    "AH100-1","BD-D75","BD-AJ112","AH100-2","BD-AJ117","BD-AJ115","ZT1","BD-Q120","BD-F38",
    "FPQ1","BD-D37","BD-Q12","BD-F37-1","BD-D69-2","DPR2","BD-H01-3","BD-D10-1","BD-D35",
    "BD-H01-10","BD-H01-9","BD-D38","DPD2","BD-G811","DPD1","DPC1","BD-H01-4","BD-LG91",
    "BD-LG10","FP16","DPC3","DPC2","DPR1","BD-H01-2","DPJ1","BD-H01-6","BD-D36","BD-H01-5",
    "BD-H01-8","AJ101","BD-H01-7","BD-H01-1","AJ102","AJ103","AJ104","FP1","FP2","FP3","FP4",
    "DPX1T","BD-L63","BD-D85","BD-ZT3T","BD-STR001","XPG1","XPG2","XPG3","XPG4","XPG5",
    "XPG6","XPS","X14","X15","BD-D39","BD-D32","BD-Q110","BD-Q130-1","BD-Q130-2","BD-F37-2",
    "AJ105","AJ106","AJ107","BD-AJ110","BD-AJ111","BD-AJ113","BD-AJ114","BD-AJ116",
    "AJ120","AJ121","AJ122","AJ123","XM1","XM2","XM3","BD-F51","BD-D69-1","BD-J47",
    "BD-J48","BD-A40D","BD-G341","BD-G812","BD-G813","BD-G814","BD-G815","BD-G816",
    "BD-G817","BD-G818","BD-G821","BD-G822","BD-G823","BD-G824","BD-G825","BD-G826",
    "BD-G827","BD-G831","BD-G832","BD-G833","BD-G834","BD-G835","BD-G836","BD-G837",
    "BD-G841","BD-G842","BD-G843","BD-G844","BD-G845","BD-G846","BD-G847","BD-GP3000",
    "BD-GP3010","BD-GP3020","BD-GP3040"
]


def init_db():
    """初始化 PostgreSQL 表结构（幂等）"""
    conn = get_db()
    cur = get_cursor(conn)
    try:
        cur.execute('''
            CREATE TABLE IF NOT EXISTS sales_records (
                id BIGSERIAL PRIMARY KEY,
                date TIMESTAMPTZ NOT NULL,
                cust TEXT NOT NULL,
                product TEXT NOT NULL,
                spec TEXT NOT NULL,
                unit TEXT DEFAULT '',
                qty REAL DEFAULT 0,
                price REAL DEFAULT 0,
                amount REAL DEFAULT 0,
                created_at TIMESTAMPTZ DEFAULT NOW(),
                UNIQUE(date, cust, product, spec, unit, qty, price, amount)
            )
        ''')
        cur.execute('''
            CREATE TABLE IF NOT EXISTS targets (
                id BIGSERIAL PRIMARY KEY,
                year INTEGER UNIQUE NOT NULL,
                amount REAL NOT NULL,
                updated_at TIMESTAMPTZ DEFAULT NOW()
            )
        ''')
        cur.execute('''
            CREATE TABLE IF NOT EXISTS new_products (
                id BIGSERIAL PRIMARY KEY,
                spec TEXT UNIQUE NOT NULL,
                created_at TIMESTAMPTZ DEFAULT NOW()
            )
        ''')

        cur.execute('''
            CREATE TABLE IF NOT EXISTS customers (
                id BIGSERIAL PRIMARY KEY,
                name TEXT UNIQUE NOT NULL,
                code TEXT,
                level CHAR(1) DEFAULT 'D',
                last_followup DATE,
                next_followup DATE,
                created_at TIMESTAMPTZ DEFAULT NOW(),
                updated_at TIMESTAMPTZ DEFAULT NOW()
            )
        ''')

        cur.execute('CREATE INDEX IF NOT EXISTS idx_customers_name ON customers(name)')
        cur.execute('CREATE INDEX IF NOT EXISTS idx_customers_level ON customers(level)')

        cur.execute("ALTER TABLE customers ADD COLUMN IF NOT EXISTS purpose_rule TEXT DEFAULT 'auto'")
        cur.execute("ALTER TABLE customers ADD COLUMN IF NOT EXISTS custom_suggestion TEXT")
        cur.execute("ALTER TABLE customers ADD COLUMN IF NOT EXISTS last_purchase_date DATE")
        cur.execute("ALTER TABLE customers ADD COLUMN IF NOT EXISTS status TEXT DEFAULT ''")
        cur.execute("ALTER TABLE customers ADD COLUMN IF NOT EXISTS note TEXT DEFAULT ''")

        # 首次初始化时写入默认新品规格
        cur.execute("SELECT COUNT(*) AS cnt FROM new_products")
        if cur.fetchone()["cnt"] == 0:
            for s in DEFAULT_NEW_PRODS:
                cur.execute(
                    "INSERT INTO new_products (spec) VALUES (%s) ON CONFLICT (spec) DO NOTHING",
                    (s,)
                )

        conn.commit()
    finally:
        cur.close()
        conn.close()


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/health')
def health():
    try:
        conn = get_db()
        cur = get_cursor(conn)
        cur.execute("SELECT 1 AS ok")
        cur.close()
        conn.close()
        return jsonify({"status": "ok", "db": "connected"})
    except Exception as e:
        return jsonify({"status": "error", "db": str(e)}), 500


# ── 数据接口 ──
@app.route('/api/data')
def get_data():
    start = request.args.get('start', '')
    end = request.args.get('end', '')
    conn = get_db()
    cur = get_cursor(conn)
    try:
        sql = "SELECT date, cust, product, spec, unit, qty, price, amount FROM sales_records"
        params = []
        conditions = []
        if start:
            conditions.append("date >= %s")
            params.append(start)
        if end:
            conditions.append("date <= %s")
            params.append(end if ' ' in end else end)
        if conditions:
            sql += " WHERE " + " AND ".join(conditions)
        sql += " ORDER BY date DESC"
        cur.execute(sql, params)
        rows = cur.fetchall()
        return jsonify([{
            "date": r["date"].strftime('%Y-%m-%d') if hasattr(r["date"], 'strftime') else str(r["date"])[:10],
            "cust": r["cust"],
            "product": r["product"],
            "spec": r["spec"],
            "unit": r["unit"],
            "qty": float(r["qty"]),
            "price": float(r["price"]),
            "amount": float(r["amount"])
        } for r in rows])
    finally:
        cur.close()
        conn.close()


@app.route('/api/summary')
def get_summary():
    start = request.args.get('start', '')
    end = request.args.get('end', '')
    prev_start = request.args.get('prev_start', '')
    prev_end = request.args.get('prev_end', '')
    year = request.args.get('year', '')

    conn = get_db()
    cur = get_cursor(conn)
    try:
        def _agg(s, e):
            """用SQL聚合计算总额和客户数"""
            e_fixed = (e if (' ' in e) else (e)) if e else None
            if s and e_fixed:
                cur.execute(
                    "SELECT COALESCE(SUM(amount),0) as total, COUNT(DISTINCT cust) as cust_cnt FROM sales_records WHERE date >= %s AND date <= %s",
                    (s, e_fixed)
                )
            elif s:
                cur.execute(
                    "SELECT COALESCE(SUM(amount),0) as total, COUNT(DISTINCT cust) as cust_cnt FROM sales_records WHERE date >= %s",
                    (s,)
                )
            elif e_fixed:
                cur.execute(
                    "SELECT COALESCE(SUM(amount),0) as total, COUNT(DISTINCT cust) as cust_cnt FROM sales_records WHERE date <= %s",
                    (e_fixed,)
                )
            else:
                cur.execute("SELECT COALESCE(SUM(amount),0) as total, COUNT(DISTINCT cust) as cust_cnt FROM sales_records")
            r = cur.fetchone()
            return float(r["total"]), r["cust_cnt"]

        cur_total, cur_cust_cnt = _agg(start, end)

        # 新客户金额：首次购买在当前时间段内的客户
        s_val = start
        e_fixed = (end if (' ' in end) else (end)) if end else None
        new_amt = 0
        old_amt = 0
        new_cust_cnt = 0
        if s_val:
            if s_val and e_fixed:
                cur.execute("""
                    SELECT COALESCE(SUM(sr.amount),0) as new_total FROM sales_records sr
                    INNER JOIN (SELECT cust, MIN(date) as first_date FROM sales_records GROUP BY cust) fo
                    ON sr.cust = fo.cust
                    WHERE fo.first_date >= %s AND fo.first_date <= %s AND sr.date >= %s AND sr.date <= %s
                """, (s_val, e_fixed, s_val, e_fixed))
            elif s_val:
                cur.execute("""
                    SELECT COALESCE(SUM(sr.amount),0) as new_total FROM sales_records sr
                    INNER JOIN (SELECT cust, MIN(date) as first_date FROM sales_records GROUP BY cust) fo
                    ON sr.cust = fo.cust
                    WHERE fo.first_date >= %s AND sr.date >= %s
                """, (s_val, s_val))
            new_amt = float(cur.fetchone()["new_total"])
            old_amt = cur_total - new_amt
            # 新客户数量
            if s_val and e_fixed:
                cur.execute("SELECT COUNT(*) as cnt FROM (SELECT cust, MIN(date) as first_date FROM sales_records GROUP BY cust) fo WHERE first_date >= %s AND first_date <= %s", (s_val, e_fixed))
            else:
                cur.execute("SELECT COUNT(*) as cnt FROM (SELECT cust, MIN(date) as first_date FROM sales_records GROUP BY cust) fo WHERE first_date >= %s", (s_val,))
            new_cust_cnt = cur.fetchone()["cnt"]
        old_cust_cnt = cur_cust_cnt - new_cust_cnt if s_val else 0

        # 新品销售额（SQL聚合）
        np_amt = 0
        cur.execute("SELECT spec FROM new_products")
        np_specs = [r["spec"] for r in cur.fetchall()]
        if np_specs and s_val and e_fixed:
            placeholders = ','.join(['%s'] * len(np_specs))
            cur.execute(
                f"SELECT COALESCE(SUM(amount),0) as total FROM sales_records WHERE spec IN ({placeholders}) AND date >= %s AND date <= %s",
                np_specs + [s_val, e_fixed]
            )
            np_amt = float(cur.fetchone()["total"])
        elif np_specs and s_val:
            placeholders = ','.join(['%s'] * len(np_specs))
            cur.execute(
                f"SELECT COALESCE(SUM(amount),0) as total FROM sales_records WHERE spec IN ({placeholders}) AND date >= %s",
                np_specs + [s_val]
            )
            np_amt = float(cur.fetchone()["total"])
        elif np_specs:
            placeholders = ','.join(['%s'] * len(np_specs))
            cur.execute(
                f"SELECT COALESCE(SUM(amount),0) as total FROM sales_records WHERE spec IN ({placeholders})",
                np_specs
            )
            np_amt = float(cur.fetchone()["total"])

        # 环比数据
        prev_total = 0
        if prev_start or prev_end:
            prev_total, _ = _agg(prev_start, prev_end)

        # 目标完成率
        target_val = 0
        if year:
            cur.execute("SELECT amount FROM targets WHERE year = %s", (int(year),))
            t = cur.fetchone()
            if t:
                target_val = float(t["amount"])

        return jsonify({
            "total_amount": cur_total,
            "customer_count": cur_cust_cnt,
            "new_customer_count": new_cust_cnt,
            "old_customer_count": old_cust_cnt,
            "new_customer_amount": new_amt,
            "old_customer_amount": old_amt,
            "new_product_amount": np_amt,
            "prev_total_amount": prev_total,
            "target_amount": target_val
        })
    finally:
        cur.close()
        conn.close()


# ── 安全转换 ──
def safe_float(val, default=0.0):
    """安全转浮点数，空值/空字符串返回默认值"""
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s == '':
        return default
    try:
        return float(s)
    except ValueError:
        return default


def normalize_date(val):
    """规范化日期为 YYYY-MM-DD HH:MM:SS 格式（兼容 PostgreSQL TIMESTAMPTZ）"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d %H:%M:%S')

    s = str(val).strip()
    if not s:
        return None

    # 处理 Excel 日期序列号（如 46175 = 2026-06-02）
    try:
        if s.isdigit() and 1 <= int(s) <= 100000:
            serial = int(s)
            excel_epoch = datetime(1899, 12, 30)
            return (excel_epoch + timedelta(days=serial)).strftime('%Y-%m-%d 00:00:00')
    except (ValueError, OverflowError):
        pass

    # 尝试带时间的格式
    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S',
                '%Y-%m-%d %H:%M', '%Y/%m/%d %H:%M',
                '%Y-%m-%dT%H:%M:%S', '%Y-%m-%dT%H:%M']:
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d %H:%M:%S')
        except ValueError:
            continue

    # 尝试纯日期格式（无时间则补 00:00:00）
    for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%m/%d/%Y', '%d/%m/%Y']:
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d 00:00:00')
        except ValueError:
            continue

    # 兜底：正则补零后再试（处理 2023/3/5 这种非补零日期）
    import re
    padded = re.sub(r'(?<=[-/.])(\d)(?=[-/.]|$)', r'0\1', s)
    for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d']:
        try:
            return datetime.strptime(padded, fmt).strftime('%Y-%m-%d 00:00:00')
        except ValueError:
            continue

    # 最终回退：返回原始值
    return s


# ── 上传接口 ──
@app.route('/api/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return jsonify({"error": "未找到文件"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "文件名为空"}), 400

    filepath = os.path.join(UPLOAD_DIR, file.filename)
    file.save(filepath)

    conn = None
    cur = None
    total_rows = 0
    skip_count = 0
    sample_empty = []
    first_headers = []
    try:
        # 流式读取，不加载全部到内存
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        # 读表头（第一行）
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if first_row is None:
            os.remove(filepath)
            return jsonify({"error": "文件为空"}), 400
        headers = [cell for cell in first_row]
        first_headers = [str(h) for h in headers[:10]]

        idx_map = {}
        col_map = {
            '终审日期': 'date', '客户名称': 'cust', '品名': 'product',
            '货品规格': 'spec', '单位': 'unit', '受订数量': 'qty',
            '单价': 'price', '本位币': 'amount'
        }
        for zh, en in col_map.items():
            for i, h in enumerate(headers):
                if h and str(h).strip() == zh:
                    idx_map[en] = i
                    break

        required = ['date', 'cust', 'product']
        for k in required:
            if k not in idx_map:
                wb.close()
                os.remove(filepath)
                return jsonify({"error": f"缺少必需列：{k}"}), 400

        conn = get_db()
        cur = get_cursor(conn)

        # 批量插入缓冲区
        batch = []
        BATCH_SIZE = 2000
        row_idx = 1  # 从第2行开始

        INSERT_SQL = '''
            INSERT INTO sales_records AS t (date, cust, product, spec, unit, qty, price, amount)
            VALUES %s
        '''

        insert_errors = []
        batch_dup_count = 0
        def flush_batch():
            nonlocal total_rows
            if not batch:
                return
            # 批次内去重：同 (date, cust, product, spec) 保留最后一条
            seen = {}
            for item in reversed(batch):
                key = (item[0], item[1], item[2], item[3], item[4], item[5], item[6], item[7])
                if key not in seen:
                    seen[key] = item
            deduped = list(reversed(list(seen.values())))
            nonlocal batch_dup_count
            batch_dup_count += len(batch) - len(deduped)
            try:
                execute_values(cur, INSERT_SQL, deduped, page_size=BATCH_SIZE)
                total_rows += len(deduped)
                conn.commit()
            except Exception as e:
                conn.rollback()
                insert_errors.append(str(e))
                print(f"[Excel Upload] 批量插入失败: {e}", flush=True)
                traceback.print_exc(file=sys.stdout)
            batch.clear()

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_idx += 1
            date_val = row[idx_map['date']] if idx_map.get('date') is not None else None
            cust_val = row[idx_map['cust']] if idx_map.get('cust') is not None else None
            product_val = row[idx_map['product']] if idx_map.get('product') is not None else None
            spec_val = row[idx_map['spec']] if idx_map.get('spec') is not None else None

            if not date_val or not cust_val or not product_val:
                skip_count += 1
                if len(sample_empty) < 3:
                    sample_empty.append({
                        "row_idx": row_idx,
                        "date": repr(date_val),
                        "cust": repr(cust_val),
                        "product": repr(product_val),
                        "spec": repr(spec_val)
                    })
                continue

            # 日期处理
            date_str = normalize_date(date_val)

            unit_val = row[idx_map.get('unit', -1)] if idx_map.get('unit') is not None and idx_map.get('unit') < len(row) and row[idx_map['unit']] is not None else ''
            if unit_val is None:
                unit_val = ''
            else:
                unit_val = str(unit_val)
            qty_val = safe_float(row[idx_map.get('qty', -1)]) if idx_map.get('qty') is not None else 0
            price_val = safe_float(row[idx_map.get('price', -1)]) if idx_map.get('price') is not None else 0
            amount_val = safe_float(row[idx_map.get('amount', -1)]) if idx_map.get('amount') is not None else 0

            batch.append((date_str, str(cust_val).strip(), str(product_val).strip(),
                         str(spec_val).strip(), unit_val, qty_val, price_val, amount_val))

            if len(batch) >= BATCH_SIZE:
                flush_batch()

        flush_batch()

        wb.close()
    except Exception as e:
        if os.path.exists(filepath):
            os.remove(filepath)
        return jsonify({"error": f"解析失败：{str(e)}"}), 400
    finally:
        if cur is not None:
            cur.close()
        if conn is not None:
            conn.close()
        if os.path.exists(filepath):
            os.remove(filepath)

    sync_customers_from_sales()
    refresh_all_customer_levels()
    refresh_all_customer_followup()

    return jsonify({
        "success": True,
        "count": total_rows,
        "total_rows_in_file": total_rows + skip_count,
        "skipped": skip_count,
        "sample_skipped": sample_empty,
        "idx_map": {k: v for k, v in idx_map.items()},
        "header_count": len(headers),
        "first_3_headers": [str(h) for h in headers[:10]],
        "insert_errors": insert_errors[:3] if insert_errors else [],
        "batch_dup_removed": batch_dup_count
    })


# ── CSV 上传接口（流式读取，适合大数据文件）──
@app.route('/api/upload-csv', methods=['POST'])
def upload_csv():
    if 'file' not in request.files:
        return jsonify({"error": "未找到文件"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "文件名为空"}), 400

    # 读取文件原始内容，尝试多种编码
    raw_bytes = file.read()
    text = None
    encoding_used = 'utf-8-sig'
    for enc in ['utf-8-sig', 'utf-8', 'gbk', 'gb2312']:
        try:
            text = raw_bytes.decode(enc)
            encoding_used = enc
            break
        except (UnicodeDecodeError, UnicodeError):
            continue

    if text is None:
        return jsonify({"error": "无法识别文件编码，请保存为 UTF-8 格式的 CSV"}), 400

    # 释放原始字节，只保留解码后的文本
    del raw_bytes

    reader = csv.reader(io.StringIO(text))
    headers_row = next(reader, None)
    if headers_row is None:
        return jsonify({"error": "文件为空"}), 400

    headers = [str(h).strip() for h in headers_row]

    # 列名映射（与 Excel 上传保持一致）
    col_map = {
        '终审日期': 'date', '客户名称': 'cust', '品名': 'product',
        '货品规格': 'spec', '单位': 'unit', '受订数量': 'qty',
        '单价': 'price', '本位币': 'amount'
    }
    idx_map = {}
    for zh, en in col_map.items():
        for i, h in enumerate(headers):
            if h == zh:
                idx_map[en] = i
                break

    required = ['date', 'cust', 'product']
    for k in required:
        if k not in idx_map:
            return jsonify({"error": f"缺少必需列：{k}（当前表头：{headers[:10]}）"}), 400

    conn = None
    cur = None
    total_rows = 0
    skip_count = 0
    sample_empty = []
    batch = []
    BATCH_SIZE = 2000

    INSERT_SQL = '''
        INSERT INTO sales_records AS t (date, cust, product, spec, unit, qty, price, amount)
        VALUES %s
        ON CONFLICT(date, cust, product, spec, unit, qty, price, amount) DO NOTHING
    '''

    insert_errors = []
    batch_dup_count = 0  # 统计批次内去重数量
    def flush_batch():
        nonlocal total_rows
        if not batch:
            return
        # 批次内去重：同 (date, cust, product, spec) 保留最后一条
        seen = {}
        for item in reversed(batch):
            key = (item[0], item[1], item[2], item[3], item[4], item[5], item[6], item[7])
            if key not in seen:
                seen[key] = item
        deduped = list(reversed(list(seen.values())))
        nonlocal batch_dup_count
        batch_dup_count += len(batch) - len(deduped)
        try:
            execute_values(cur, INSERT_SQL, deduped, page_size=BATCH_SIZE)
            total_rows += len(deduped)
            conn.commit()
        except Exception as e:
            conn.rollback()
            insert_errors.append(str(e))
            print(f"[CSV Upload] 批量插入失败: {e}", flush=True)
            traceback.print_exc(file=sys.stdout)
        batch.clear()

    try:
        conn = get_db()
        cur = get_cursor(conn)

        for row_idx, row in enumerate(reader, start=2):
            # 跳过空行
            if not row or all(not str(c).strip() for c in row):
                skip_count += 1
                continue

            # 安全检查：确保行有足够的列
            max_idx = max(idx_map.values())
            if len(row) <= max_idx:
                skip_count += 1
                if len(sample_empty) < 3:
                    sample_empty.append({
                        "row_idx": row_idx,
                        "reason": f"列数不足（需要{max_idx+1}列，实际{len(row)}列）"
                    })
                continue

            date_val = row[idx_map['date']].strip() if idx_map.get('date') is not None else ''
            cust_val = row[idx_map['cust']].strip() if idx_map.get('cust') is not None else ''
            product_val = row[idx_map['product']].strip() if idx_map.get('product') is not None else ''
            spec_val = row[idx_map['spec']].strip() if idx_map.get('spec') is not None else ''

            if not date_val or not cust_val or not product_val:
                skip_count += 1
                if len(sample_empty) < 3:
                    sample_empty.append({
                        "row_idx": row_idx,
                        "date": repr(date_val),
                        "cust": repr(cust_val),
                        "product": repr(product_val),
                        "spec": repr(spec_val)
                    })
                continue

            # 日期格式处理
            date_str = normalize_date(date_val)

            unit_val = row[idx_map['unit']].strip() if idx_map.get('unit') is not None and idx_map['unit'] < len(row) else ''
            qty_val = safe_float(row[idx_map['qty']]) if idx_map.get('qty') is not None and idx_map['qty'] < len(row) else 0
            price_val = safe_float(row[idx_map['price']]) if idx_map.get('price') is not None and idx_map['price'] < len(row) else 0
            amount_val = safe_float(row[idx_map['amount']]) if idx_map.get('amount') is not None and idx_map['amount'] < len(row) else 0

            batch.append((date_str, cust_val, product_val, spec_val,
                         unit_val, qty_val, price_val, amount_val))

            if len(batch) >= BATCH_SIZE:
                flush_batch()

        flush_batch()

    except Exception as e:
        return jsonify({"error": f"解析失败：{str(e)}"}), 400
    finally:
        if cur is not None:
            cur.close()
        if conn is not None:
            conn.close()

    sync_customers_from_sales()
    refresh_all_customer_levels()
    refresh_all_customer_followup()

    return jsonify({
        "success": True,
        "count": total_rows,
        "total_rows_in_file": total_rows + skip_count,
        "skipped": skip_count,
        "sample_skipped": sample_empty,
        "encoding": encoding_used,
        "header_count": len(headers),
        "headers": headers[:10],
        "insert_errors": insert_errors[:3] if insert_errors else [],
        "batch_dup_removed": batch_dup_count
    })


# ── 目标接口 ──
@app.route('/api/target', methods=['GET', 'POST', 'DELETE'])
def target():
    conn = get_db()
    cur = get_cursor(conn)
    try:
        if request.method == 'GET':
            year = request.args.get('year', '')
            if year:
                cur.execute("SELECT year, amount FROM targets WHERE year = %s", (int(year),))
                row = cur.fetchone()
                if row:
                    return jsonify({"year": row["year"], "amount": float(row["amount"])})
                return jsonify({"year": int(year), "amount": 0})
            else:
                cur.execute("SELECT year, amount FROM targets ORDER BY year")
                rows = cur.fetchall()
                return jsonify({str(r["year"]): float(r["amount"]) for r in rows})

        elif request.method == 'POST':
            data = request.get_json()
            year = int(data.get("year", 0))
            amount = float(data.get("amount", 0))
            if not year or amount <= 0:
                return jsonify({"error": "无效参数"}), 400
            cur.execute(
                "INSERT INTO targets (year, amount, updated_at) VALUES (%s, %s, NOW()) "
                "ON CONFLICT(year) DO UPDATE SET amount = EXCLUDED.amount, updated_at = NOW()",
                (year, amount)
            )
            conn.commit()
            return jsonify({"success": True, "year": year, "amount": amount})

        elif request.method == 'DELETE':
            year = request.args.get('year', '')
            if year:
                cur.execute("DELETE FROM targets WHERE year = %s", (int(year),))
                conn.commit()
            return jsonify({"success": True})
    finally:
        cur.close()
        conn.close()


# ── 新品规格接口 ──
@app.route('/api/new-products', methods=['GET', 'POST'])
def new_products():
    conn = get_db()
    cur = get_cursor(conn)
    try:
        if request.method == 'GET':
            cur.execute("SELECT id, spec FROM new_products ORDER BY id")
            rows = cur.fetchall()
            return jsonify([{"id": r["id"], "spec": r["spec"]} for r in rows])

        elif request.method == 'POST':
            data = request.get_json()
            specs = data.get("specs", [])
            if isinstance(specs, list):
                cur.execute("DELETE FROM new_products")
                for s in specs:
                    if s and str(s).strip():
                        cur.execute(
                            "INSERT INTO new_products (spec) VALUES (%s) ON CONFLICT DO NOTHING",
                            (str(s).strip(),)
                        )
                conn.commit()
            return jsonify({"success": True, "count": len(specs)})
    finally:
        cur.close()
        conn.close()


@app.route('/api/new-products/<int:idx>', methods=['DELETE'])
def delete_new_product(idx):
    conn = get_db()
    cur = get_cursor(conn)
    try:
        cur.execute("DELETE FROM new_products WHERE id = %s", (idx,))
        conn.commit()
        return jsonify({"success": True})
    finally:
        cur.close()
        conn.close()


# ==================== 客户管理与跟进任务模块 ====================

FOLLOWUP_DAYS = {
    'A': 7, 'B': 14,
    'C1': 14, 'C2': 30, 'C3': 60, 'C4': 90, 'C': 30,  # C defaults to 30
    'D1': 30, 'D2': 90, 'D3': 0, 'D4': 0, 'D': 30     # D3/D4 no active followup
}
RECOVERY_DAYS = 60


def get_customer_year_sales(cust_name, year):
    """获取客户指定年份的销售额"""
    conn = get_db()
    cur = get_cursor(conn)
    try:
        start = f"{year}-01-01"
        end = f"{year}-12-31 23:59:59"
        cur.execute(
            "SELECT COALESCE(SUM(amount), 0) as total FROM sales_records WHERE cust = %s AND date >= %s AND date <= %s",
            (cust_name, start, end)
        )
        result = cur.fetchone()
        return float(result['total']) if result else 0
    finally:
        cur.close()
        conn.close()


def get_customer_level_by_sales(sales):
    """根据销售额返回等级"""
    if sales >= 50000:
        return 'A'
    elif sales >= 10000:
        return 'B'
    elif sales >= 2000:
        return 'C'
    else:
        return 'D'


# 子等级缓存（请求级，避免重复查询）
_sub_level_cache = {}

def get_customer_sub_level(cust_name, level):
    """获取C/D类客户的子标签（C1-C4, D1-D4），优先用缓存"""
    if level not in ('C', 'D'):
        return level, level, FOLLOWUP_DAYS.get(level, 30)
    # 缓存命中
    cache_key = cust_name
    if cache_key in _sub_level_cache:
        return _sub_level_cache[cache_key]
    # 未命中，返回默认（批量查询后再填充）
    return level, level, FOLLOWUP_DAYS.get(level, 30)


def preload_all_sub_levels():
    """一次性载入所有 C/D 客户的子标签（1条SQL），存入缓存"""
    global _sub_level_cache
    _sub_level_cache = {}
    from datetime import date
    today = date.today()
    conn = get_db()
    cur = get_cursor(conn)
    try:
        cur.execute("""
            WITH cust_stats AS (
                SELECT c.name, c.level, c.status, c.note, c.created_at,
                    COUNT(s.amount) as total_orders,
                    COALESCE(MIN(s.date), NULL) as first_order_date,
                    COALESCE(MAX(s.date), NULL) as last_order_date,
                    COALESCE(MIN(s.amount), 0) as first_amount
                FROM customers c
                LEFT JOIN sales_records s ON c.name = s.cust
                GROUP BY c.name, c.level, c.status, c.note, c.created_at
            )
            SELECT name, level, total_orders, first_amount, last_order_date, status, note, created_at
            FROM cust_stats
        """)
        rows = cur.fetchall()
        for r in rows:
            name = r['name']
            level = r['level']
            if level not in ('C', 'D'):
                _sub_level_cache[name] = (level, level, FOLLOWUP_DAYS.get(level, 30))
                continue

            if level == 'C':
                total_orders = int(r['total_orders'])
                first_amount = float(r['first_amount']) if r['first_amount'] else 0
                last_order = r['last_order_date']
                days_since = (today - last_order.date()).days if last_order and hasattr(last_order, 'date') else 999

                if total_orders == 1 and first_amount < 500:
                    sub, label = 'C4', '一次性客户'
                elif days_since <= 90:
                    sub, label = 'C1', '高频活跃客户'
                elif days_since <= 180:
                    sub, label = 'C2', '低频客户'
                elif days_since <= 365:
                    sub, label = 'C3', '沉睡客户'
                else:
                    sub, label = 'C3', '沉睡客户'
            else:  # D
                total_orders = int(r['total_orders'])
                has_inquiry = (r['status'] == 'inquiry') or (r['note'] and '询价' in str(r['note']))
                recently_added = r['created_at'] and (today - r['created_at'].date()).days <= 30

                if has_inquiry or recently_added:
                    sub, label = 'D1', '新线索客户'
                elif total_orders >= 1:
                    sub, label = 'D2', '流失老客户'
                elif r['created_at'] is not None:
                    sub, label = 'D3', '无效线索'
                else:
                    sub, label = 'D4', '从未互动客户'

            _sub_level_cache[name] = (sub, label, FOLLOWUP_DAYS.get(sub, 30))
    finally:
        cur.close()
        conn.close()
    """获取客户近3年内的最高等级"""
    current_year = datetime.now().year
    years = [current_year - 2, current_year - 1, current_year]
    peak = 'D'
    level_order = {'A': 4, 'B': 3, 'C': 2, 'D': 1}
    for y in years:
        sales = get_customer_year_sales(cust_name, y)
        level = get_customer_level_by_sales(sales)
        if level_order.get(level, 0) > level_order.get(peak, 0):
            peak = level
    return peak


def refresh_all_customer_levels():
    """刷新所有客户的等级（基于上年销售额）"""
    conn = get_db()
    cur = get_cursor(conn)
    try:
        last_year = datetime.now().year - 1
        cur.execute("SELECT DISTINCT cust FROM sales_records")
        customers = cur.fetchall()
        for row in customers:
            cust_name = row['cust']
            sales = get_customer_year_sales(cust_name, last_year)
            level = get_customer_level_by_sales(sales)
            cur.execute(
                "UPDATE customers SET level = %s, updated_at = NOW() WHERE name = %s",
                (level, cust_name)
            )
        conn.commit()
    finally:
        cur.close()
        conn.close()


def calculate_next_followup(level, peak_level, last_followup):
    """计算下次跟进日期"""
    from datetime import date, timedelta
    today = date.today()
    if last_followup is None:
        return today  # Will be updated by caller with last_purchase_date
    base_days = FOLLOWUP_DAYS.get(level, 30)
    next_date = last_followup + timedelta(days=base_days)
    if peak_level in ['A', 'B'] and level in ['C', 'D']:
        recovery_date = last_followup + timedelta(days=RECOVERY_DAYS)
        if recovery_date < next_date:
            next_date = recovery_date
    if next_date < today:
        next_date = today
    return next_date


def refresh_all_customer_followup():
    """刷新所有客户的下次跟进日期"""
    conn = get_db()
    cur = get_cursor(conn)
    try:
        cur.execute("SELECT name, level, last_followup FROM customers")
        customers = cur.fetchall()
        for row in customers:
            cust_name = row['name']
            level = row['level']
            last_followup = row['last_followup']
            peak_level = get_customer_peak_level(cust_name)
            next_followup = calculate_next_followup(level, peak_level, last_followup)
            cur.execute(
                "UPDATE customers SET next_followup = %s, updated_at = NOW() WHERE name = %s",
                (next_followup, cust_name)
            )
        conn.commit()
    finally:
        cur.close()
        conn.close()


def sync_customers_from_sales():
    """从销售记录同步客户到customers表"""
    conn = get_db()
    cur = get_cursor(conn)
    try:
        cur.execute("SELECT DISTINCT cust FROM sales_records")
        rows = cur.fetchall()
        for row in rows:
            cust_name = row['cust']
            cur.execute(
                "INSERT INTO customers (name, code, level, last_followup, next_followup) VALUES (%s, NULL, 'D', NULL, NULL) ON CONFLICT (name) DO NOTHING",
                (cust_name,)
            )
        conn.commit()
    finally:
        cur.close()
        conn.close()


def get_task_purpose_and_suggestion(cust_name, level, peak_level, last_order_date, last_followup, year_sales, consecutive_decline=False):
    """
    根据客户数据判断任务目的和行动建议
    返回: {'purpose': str, 'suggestion': str, 'priority': int}
    """
    from datetime import date, timedelta, datetime as dt
    today = date.today()
    # Normalize to date
    lod = last_order_date.date() if hasattr(last_order_date, 'date') else (last_order_date if isinstance(last_order_date, date) else None)
    lfu = last_followup.date() if hasattr(last_followup, 'date') else (last_followup if isinstance(last_followup, date) else None)
    days_since_last_order = (today - lod).days if lod else 999
    days_since_followup = (today - lfu).days if lfu else 999
    days_overdue = (today - lfu).days if lfu and lfu < today else 0

    # 1. 逾期超过7天 且 A/B类
    if days_overdue > 7 and level in ('A', 'B'):
        return {'purpose': '紧急跟进', 'suggestion': '客户已逾期一周，建议立即电话联系，了解是否存在问题', 'priority': 1}
    # 2. 逾期超过3天
    if days_overdue > 3:
        return {'purpose': '逾期提醒', 'suggestion': '已逾期，建议尽快联系客户', 'priority': 2}
    # 3. A类 且 30天内有订单
    if level == 'A' and days_since_last_order <= 30:
        return {'purpose': '客情维护', 'suggestion': '了解使用情况，推送新品信息，保持良好的客户关系', 'priority': 3}
    # 4. A类 且 超过60天无订单
    if level == 'A' and days_since_last_order > 60:
        return {'purpose': '唤醒', 'suggestion': '客户长时间未下单，了解是否被竞争对手抢占，邀请参加新品发布会', 'priority': 4}
    # 5. C/D 且 历史为A/B
    if level in ('C', 'D') and peak_level in ('A', 'B'):
        return {'purpose': '挽回', 'suggestion': '历史大客户已流失，发送优惠券或新品资料，了解流失原因', 'priority': 5}
    # 6. 连续销售额下降
    if consecutive_decline:
        return {'purpose': '风险预警', 'suggestion': '销售额连续下降，了解客户经营状况，是否需要支持', 'priority': 6}
    # 7. 新客户（首次成交后14天内）
    if days_since_last_order <= 14 and (year_sales or 0) > 0:
        return {'purpose': '回访', 'suggestion': '新客户首次购买后回访，询问使用体验，收集反馈', 'priority': 7}
    # 8. 默认
    return {'purpose': '常规跟进', 'suggestion': '了解近期需求，推送产品资料，保持联系', 'priority': 8}


# ==================== 新增API接口 ====================

@app.route('/api/customers/stats')
def get_customer_stats():
    """获取各等级客户数量（先按客户聚合再按等级计数）"""
    year = request.args.get('year', '')
    if not year:
        year = str(datetime.now().year - 1)
    else:
        year = str(int(year))
    conn = get_db()
    cur = get_cursor(conn)
    try:
        cur.execute("""
            SELECT level, COUNT(*) as cnt FROM (
                SELECT c.name,
                    CASE
                        WHEN COALESCE(SUM(s.amount), 0) >= 50000 THEN 'A'
                        WHEN COALESCE(SUM(s.amount), 0) >= 10000 THEN 'B'
                        WHEN COALESCE(SUM(s.amount), 0) >= 2000 THEN 'C'
                        ELSE 'D'
                    END as level
                FROM customers c
                LEFT JOIN sales_records s ON c.name = s.cust
                    AND s.date >= %s AND s.date <= %s
                GROUP BY c.name
            ) sub
            GROUP BY level
            ORDER BY CASE level WHEN 'A' THEN 1 WHEN 'B' THEN 2 WHEN 'C' THEN 3 ELSE 4 END
        """, (f"{year}-01-01", f"{year}-12-31 23:59:59"))
        stats = {'A': 0, 'B': 0, 'C': 0, 'D': 0}
        for row in cur.fetchall():
            stats[row['level']] = row['cnt']
        return jsonify({'success': True, 'data': stats, 'year': year})
    finally:
        cur.close()
        conn.close()


@app.route('/api/customers/by-level')
def get_customers_by_level():
    """获取某等级客户列表（分页，SQL JOIN）"""
    level = request.args.get('level', '')
    year = request.args.get('year', '')
    page = int(request.args.get('page', 1))
    limit = int(request.args.get('limit', 20))
    if not year:
        year = str(datetime.now().year - 1)
    if level not in ['A', 'B', 'C', 'D']:
        return jsonify({'error': '无效等级'}), 400
    conn = get_db()
    cur = get_cursor(conn)
    try:
        cur.execute("""
            WITH cust_sales AS (
                SELECT c.name, c.code, COALESCE(SUM(s.amount), 0) as total
                FROM customers c
                LEFT JOIN sales_records s ON c.name = s.cust
                    AND s.date >= %s AND s.date <= %s
                GROUP BY c.name, c.code
            )
            SELECT name, code, total FROM cust_sales
            WHERE
                CASE
                    WHEN total >= 50000 THEN 'A'
                    WHEN total >= 10000 THEN 'B'
                    WHEN total >= 2000 THEN 'C'
                    ELSE 'D'
                END = %s
            ORDER BY total DESC
            LIMIT %s OFFSET %s
        """, (f"{year}-01-01", f"{year}-12-31 23:59:59", level, limit, (page - 1) * limit))
        rows = cur.fetchall()
        data = [{'name': r['name'], 'code': r['code'] or '', 'sales': float(r['total'])} for r in rows]
        cur.execute("""
            SELECT COUNT(*) as cnt FROM (
                SELECT
                    CASE
                        WHEN COALESCE(SUM(s.amount), 0) >= 50000 THEN 'A'
                        WHEN COALESCE(SUM(s.amount), 0) >= 10000 THEN 'B'
                        WHEN COALESCE(SUM(s.amount), 0) >= 2000 THEN 'C'
                        ELSE 'D'
                    END as lv
                FROM customers c
                LEFT JOIN sales_records s ON c.name = s.cust
                    AND s.date >= %s AND s.date <= %s
                GROUP BY c.name
            ) sub WHERE lv = %s
        """, (f"{year}-01-01", f"{year}-12-31 23:59:59", level))
        total = cur.fetchone()['cnt']
        return jsonify({'success': True, 'data': data, 'total': total, 'page': page, 'limit': limit, 'year': year})
    finally:
        cur.close()
        conn.close()


@app.route('/api/customers/followup/today')
def get_today_tasks():
    """获取今日需跟进的客户（含任务目的和建议）"""
    from datetime import date
    today = date.today()
    y0, y1, y2 = today.year, today.year - 1, today.year - 2
    conn = get_db()
    # 预加载子等级（复用同一连接）
    global _sub_level_cache
    _sub_level_cache = {}
    try:
        cur2 = get_cursor(conn)
        cur2.execute("""
            WITH cust_stats AS (
                SELECT c.name, c.level, c.status, c.note, c.created_at,
                    COUNT(s.amount) as total_orders,
                    COALESCE(MAX(s.date), NULL) as last_order_date,
                    COALESCE(MIN(s.amount), 0) as first_amount
                FROM customers c LEFT JOIN sales_records s ON c.name = s.cust
                GROUP BY c.name, c.level, c.status, c.note, c.created_at
            )
            SELECT name, level, total_orders, first_amount, last_order_date, status, note, created_at
            FROM cust_stats
        """)
        for r in cur2.fetchall():
            nm = r['name']; lv = r['level']
            if lv not in ('C', 'D'):
                _sub_level_cache[nm] = (lv, lv, FOLLOWUP_DAYS.get(lv, 30))
            elif lv == 'C':
                orders = int(r['total_orders']); fa = float(r['first_amount'] or 0)
                lo = r['last_order_date']
                days = (today - lo.date()).days if lo and hasattr(lo, 'date') else 999
                if orders == 1 and fa < 500: sub, lbl = 'C4', '一次性客户'
                elif days <= 90: sub, lbl = 'C1', '高频活跃客户'
                elif days <= 180: sub, lbl = 'C2', '低频客户'
                else: sub, lbl = 'C3', '沉睡客户'
                _sub_level_cache[nm] = (sub, lbl, FOLLOWUP_DAYS.get(sub, 30))
            else:
                orders = int(r['total_orders'])
                has_inq = (r['status'] == 'inquiry') or (r['note'] and '询价' in str(r['note']))
                recent = r['created_at'] and (today - r['created_at'].date()).days <= 30
                if has_inq or recent: sub, lbl = 'D1', '新线索客户'
                elif orders >= 1: sub, lbl = 'D2', '流失老客户'
                elif r['created_at']: sub, lbl = 'D3', '无效线索'
                else: sub, lbl = 'D4', '从未互动客户'
                _sub_level_cache[nm] = (sub, lbl, FOLLOWUP_DAYS.get(sub, 30))
        cur2.close()
    except Exception:
        pass  # Fallback: get_customer_sub_level will return defaults
    cur = get_cursor(conn)
    try:
        cur.execute("""
            WITH yearly AS (
                SELECT c.name,
                    COALESCE(SUM(CASE WHEN s.date >= %s THEN s.amount ELSE 0 END), 0) as s0,
                    COALESCE(SUM(CASE WHEN s.date >= %s AND s.date < %s THEN s.amount ELSE 0 END), 0) as s1,
                    COALESCE(SUM(CASE WHEN s.date >= %s AND s.date < %s THEN s.amount ELSE 0 END), 0) as s2
                FROM customers c
                LEFT JOIN sales_records s ON c.name = s.cust
                GROUP BY c.name
            ),
            peak AS (
                SELECT name,
                    CASE
                        WHEN s0 >= 50000 OR s1 >= 50000 OR s2 >= 50000 THEN 'A'
                        WHEN s0 >= 10000 OR s1 >= 10000 OR s2 >= 10000 THEN 'B'
                        WHEN s0 >= 2000 OR s1 >= 2000 OR s2 >= 2000 THEN 'C'
                        ELSE 'D'
                    END as peak_level,
                    (s1 < s2 AND s0 < s1 AND s1 > 0) as is_declining
                FROM yearly
            ),
            last_order AS (
                SELECT cust, MAX(date) as last_date
                FROM sales_records
                GROUP BY cust
            )
            SELECT cu.name, cu.level, cu.last_followup, cu.next_followup,
                   p.peak_level, p.is_declining, cu.purpose_rule, cu.custom_suggestion,
                   lo.last_date as last_order_date
            FROM customers cu
            LEFT JOIN peak p ON cu.name = p.name
            LEFT JOIN last_order lo ON cu.name = lo.cust
            WHERE cu.next_followup <= %s
            ORDER BY cu.next_followup ASC,
                CASE cu.level WHEN 'A' THEN 1 WHEN 'B' THEN 2 WHEN 'C' THEN 3 ELSE 4 END
        """, (f"{y0}-01-01", f"{y1}-01-01", f"{y0}-01-01", f"{y2}-01-01", f"{y1}-01-01", today))
        tasks = cur.fetchall()
        result = []
        for r in tasks:
            last_order_date = r['last_order_date'] if r['last_order_date'] else None
            last_followup = r['last_followup']
            level = r['level']
            peak_level = r['peak_level'] or 'D'
            # Use auto rule if set to auto, or if custom_suggestion is empty
            use_auto = (r['purpose_rule'] != 'manual') or (not r['custom_suggestion'])
            if use_auto:
                ps = get_task_purpose_and_suggestion(
                    r['name'], level, peak_level,
                    last_order_date, last_followup,
                    None, r['is_declining']
                )
                purpose = ps['purpose']
                suggestion = ps['suggestion']
                priority = ps['priority']
            else:
                purpose = '自定义'
                suggestion = r['custom_suggestion'] or ''
                priority = 5
            # Get sub-level for C/D
            sub_level, sub_name, sub_freq = get_customer_sub_level(r['name'], level)
            result.append({
                'name': r['name'],
                'level': level,
                'sub_level': sub_level,
                'sub_name': sub_name,
                'peak_level': peak_level,
                'last_followup': last_followup.isoformat() if last_followup else None,
                'next_followup': r['next_followup'].isoformat() if r['next_followup'] else None,
                'is_overdue': r['next_followup'] < today if r['next_followup'] else False,
                'purpose': purpose,
                'suggestion': suggestion,
                'priority': priority,
                'purpose_rule': r['purpose_rule'] or 'auto',
                'custom_suggestion': r['custom_suggestion'] or '',
                'last_order_date': last_order_date.strftime('%Y-%m-%d') if last_order_date else None,
                'consecutive_decline': r['is_declining']
            })
        return jsonify({'success': True, 'data': result})
    finally:
        cur.close()
        conn.close()


@app.route('/api/customers/followup/rule-detail')
def get_rule_detail():
    """返回某个客户的计算详情（用于前端浮层展示）"""
    from datetime import date
    cust_name = request.args.get('customer_name', '')
    if not cust_name:
        return jsonify({'error': '缺少客户名称'}), 400
    conn = get_db()
    cur = get_cursor(conn)
    try:
        cur.execute("SELECT name, level, last_followup, next_followup, purpose_rule, custom_suggestion FROM customers WHERE name = %s", (cust_name,))
        cust = cur.fetchone()
        if not cust:
            return jsonify({'error': '客户不存在'}), 404
        peak = get_customer_peak_level(cust_name)
        today = date.today()
        level = cust['level']
        last_fu = cust['last_followup']
        next_fu = cust['next_followup']
        freq = FOLLOWUP_DAYS.get(level, 30)
        is_recovery = peak in ('A', 'B') and level in ('C', 'D')
        recovery_interval = RECOVERY_DAYS if is_recovery else None
        effective_interval = recovery_interval or freq
        is_overdue = bool(next_fu and next_fu < today)

        detail = {
            'customer_name': cust_name,
            'current_level': level,
            'peak_level': peak,
            'followup_frequency': freq,
            'is_recovery': is_recovery,
            'recovery_interval': recovery_interval,
            'last_followup_date': last_fu.isoformat() if last_fu else None,
            'next_followup_date': next_fu.isoformat() if next_fu else None,
            'today': today.isoformat(),
            'is_overdue': is_overdue,
            'effective_interval': effective_interval,
            'judgment': '逾期跟进' if is_overdue else ('挽回任务' if is_recovery else '常规跟进')
        }
        return jsonify({'success': True, 'data': detail})
    finally:
        cur.close()
        conn.close()


@app.route('/api/customers/followup/today-stats')
def get_today_stats():
    """获取今日统计：已完成数、逾期总数"""
    from datetime import date
    today = date.today()
    conn = get_db()
    cur = get_cursor(conn)
    try:
        # 逾期总数
        cur.execute("SELECT COUNT(*) as cnt FROM customers WHERE next_followup < %s", (today,))
        overdue_total = cur.fetchone()['cnt']
        # 今日待跟进总数
        cur.execute("SELECT COUNT(*) as cnt FROM customers WHERE next_followup <= %s", (today,))
        today_total = cur.fetchone()['cnt']
        return jsonify({'success': True, 'data': {
            'overdue_total': overdue_total,
            'today_total': today_total
        }})
    finally:
        cur.close()
        conn.close()


@app.route('/api/customers/followup/week')
def get_week_tasks():
    """获取本周需跟进的客户"""
    from datetime import date, timedelta
    today = date.today()
    week_end = today + timedelta(days=7)
    conn = get_db()
    cur = get_cursor(conn)
    try:
        cur.execute(
            "SELECT name, level, last_followup, next_followup FROM customers WHERE next_followup <= %s AND next_followup > %s ORDER BY next_followup ASC",
            (week_end, today)
        )
        tasks = cur.fetchall()
        result = []
        for row in tasks:
            result.append({
                'name': row['name'],
                'level': row['level'],
                'last_followup': row['last_followup'].isoformat() if row['last_followup'] else None,
                'next_followup': row['next_followup'].isoformat() if row['next_followup'] else None
            })
        return jsonify({'success': True, 'data': result})
    finally:
        cur.close()
        conn.close()


@app.route('/api/customers/followup/mark', methods=['POST'])
def mark_followup():
    """标记客户已跟进"""
    from datetime import date
    data = request.get_json()
    cust_name = data.get('customer_name', '')
    if not cust_name:
        return jsonify({'error': '缺少客户名称'}), 400
    conn = get_db()
    cur = get_cursor(conn)
    try:
        cur.execute("SELECT level FROM customers WHERE name = %s", (cust_name,))
        row = cur.fetchone()
        if not row:
            return jsonify({'error': '客户不存在'}), 404
        today = date.today()
        peak_level = get_customer_peak_level(cust_name)
        next_followup = calculate_next_followup(row['level'], peak_level, today)
        cur.execute(
            "UPDATE customers SET last_followup = %s, next_followup = %s, updated_at = NOW() WHERE name = %s",
            (today, next_followup, cust_name)
        )
        conn.commit()
        return jsonify({'success': True, 'message': '已标记跟进', 'next_followup': next_followup.isoformat()})
    finally:
        cur.close()
        conn.close()


@app.route('/api/customers/followup/batch', methods=['POST'])
def batch_mark_followup():
    """批量标记跟进"""
    from datetime import date
    data = request.get_json()
    level = data.get('level', '')
    conn = get_db()
    cur = get_cursor(conn)
    try:
        today = date.today()
        if level and level in ['A', 'B', 'C', 'D']:
            cur.execute("SELECT name, level FROM customers WHERE level = %s", (level,))
        else:
            cur.execute("SELECT name, level FROM customers WHERE next_followup <= %s", (today,))
        customers = cur.fetchall()
        count = 0
        for row in customers:
            peak_level = get_customer_peak_level(row['name'])
            next_followup = calculate_next_followup(row['level'], peak_level, today)
            cur.execute(
                "UPDATE customers SET last_followup = %s, next_followup = %s, updated_at = NOW() WHERE name = %s",
                (today, next_followup, row['name'])
            )
            count += 1
        conn.commit()
        return jsonify({'success': True, 'message': f'已批量标记{count}个客户'})
    finally:
        cur.close()
        conn.close()


@app.route('/api/customers/followup/postpone', methods=['POST'])
def postpone_followup():
    """改期：将客户的跟进日期改为指定日期"""
    data = request.get_json()
    cust_name = data.get('customer_name')
    new_date = data.get('new_followup_date')
    if not cust_name or not new_date:
        return jsonify({'error': '缺少参数'}), 400
    conn = get_db()
    cur = get_cursor(conn)
    try:
        cur.execute(
            "UPDATE customers SET next_followup = %s, updated_at = NOW() WHERE name = %s",
            (new_date, cust_name)
        )
        conn.commit()
        return jsonify({'success': True, 'message': f'已改期至 {new_date}'})
    finally:
        cur.close()
        conn.close()


@app.route('/api/customers/followup/update-suggestion', methods=['POST'])
def update_customer_suggestion():
    """更新客户的自定义建议话术"""
    data = request.get_json()
    cust_name = data.get('customer_name')
    custom_suggestion = data.get('custom_suggestion', '')
    purpose_rule = data.get('purpose_rule', 'manual')
    if not cust_name:
        return jsonify({'error': '缺少客户名称'}), 400
    conn = get_db()
    cur = get_cursor(conn)
    try:
        cur.execute(
            "UPDATE customers SET custom_suggestion = %s, purpose_rule = %s, updated_at = NOW() WHERE name = %s",
            (custom_suggestion, purpose_rule, cust_name)
        )
        conn.commit()
        return jsonify({'success': True, 'message': '建议话术已保存'})
    finally:
        cur.close()
        conn.close()


@app.route('/api/customers/update', methods=['POST'])
def update_customer():
    """更新客户名称或编码"""
    data = request.get_json()
    old_name = data.get('old_name', '')
    new_name = data.get('new_name', '')
    new_code = data.get('new_code', '')
    if not old_name:
        return jsonify({'error': '缺少原客户名称'}), 400
    conn = get_db()
    cur = get_cursor(conn)
    try:
        if new_name and new_name != old_name:
            cur.execute("UPDATE customers SET name = %s, updated_at = NOW() WHERE name = %s", (new_name, old_name))
        if new_code is not None:
            cur.execute("UPDATE customers SET code = %s, updated_at = NOW() WHERE name = %s", (new_code, new_name or old_name))
        conn.commit()
        return jsonify({'success': True, 'message': '已更新'})
    finally:
        cur.close()
        conn.close()


@app.route('/api/customers/delete', methods=['POST'])
def delete_customer():
    """删除客户"""
    data = request.get_json()
    cust_name = data.get('customer_name', '')
    if not cust_name:
        return jsonify({'error': '缺少客户名称'}), 400
    conn = get_db()
    cur = get_cursor(conn)
    try:
        cur.execute("DELETE FROM customers WHERE name = %s", (cust_name,))
        conn.commit()
        return jsonify({'success': True, 'message': '已删除'})
    finally:
        cur.close()
        conn.close()


@app.route('/api/customers/note', methods=['POST'])
def add_customer_note():
    """添加客户备注（近期报价等）"""
    data = request.get_json()
    cust_name = data.get('customer_name', '')
    note = data.get('note', '')
    status = data.get('status', '')
    if not cust_name:
        return jsonify({'error': '缺少客户名称'}), 400
    conn = get_db()
    cur = get_cursor(conn)
    try:
        if status:
            cur.execute("UPDATE customers SET status = %s, updated_at = NOW() WHERE name = %s", (status, cust_name))
        if note:
            cur.execute("UPDATE customers SET note = %s, updated_at = NOW() WHERE name = %s", (note, cust_name))
        conn.commit()
        return jsonify({'success': True, 'message': '已更新'})
    finally:
        cur.close()
        conn.close()


@app.route('/api/customers/all')
def get_all_customers():
    """获取全部客户列表（支持分页），按编码降序"""
    page = int(request.args.get('page', 1))
    limit = int(request.args.get('limit', 50))
    search = request.args.get('search', '')
    conn = get_db()
    cur = get_cursor(conn)
    try:
        if search:
            cur.execute(
                "SELECT COUNT(*) as cnt FROM customers WHERE name ILIKE %s OR code ILIKE %s",
                (f'%{search}%', f'%{search}%')
            )
        else:
            cur.execute("SELECT COUNT(*) as cnt FROM customers")
        total = cur.fetchone()['cnt']

        offset = (page - 1) * limit
        if search:
            cur.execute(
                "SELECT name, code, level FROM customers WHERE name ILIKE %s OR code ILIKE %s ORDER BY code DESC NULLS LAST, name ASC LIMIT %s OFFSET %s",
                (f'%{search}%', f'%{search}%', limit, offset)
            )
        else:
            cur.execute(
                "SELECT name, code, level FROM customers ORDER BY code DESC NULLS LAST, name ASC LIMIT %s OFFSET %s",
                (limit, offset)
            )
        rows = cur.fetchall()
        data = [{'name': r['name'], 'code': r['code'] or '', 'level': r['level']} for r in rows]
        return jsonify({'success': True, 'data': data, 'total': total, 'page': page, 'limit': limit})
    finally:
        cur.close()
        conn.close()


@app.route('/api/customers/import-codes', methods=['POST'])
def import_codes():
    """导入编码对照表（Excel）— 自动识别表头列名"""
    if 'file' not in request.files:
        return jsonify({'error': '未找到文件'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '文件名为空'}), 400
    import openpyxl
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active
    conn = get_db()
    cur = get_cursor(conn)
    try:
        # 读取表头，自动识别"名称"和"编码"列位置
        headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
        name_idx = None
        code_idx = None
        for i, h in enumerate(headers):
            if '名称' in h or '客户' in h:
                if name_idx is None or '名称' in h:
                    name_idx = i
            if '编码' in h or '代号' in h or 'code' in h.lower():
                if code_idx is None or '编码' in h:
                    code_idx = i

        # 回退：如果没找到名称列，用最后一列；编码列用倒数第二列
        if name_idx is None and code_idx is None and len(headers) >= 2:
            # 最后一列是名称，倒数第二列是编码
            name_idx = len(headers) - 1
            code_idx = len(headers) - 2
        elif name_idx is None and code_idx is not None:
            name_idx = code_idx + 1 if code_idx + 1 < len(headers) else code_idx - 1
        elif code_idx is None and name_idx is not None:
            code_idx = name_idx - 1 if name_idx > 0 else name_idx + 1

        if name_idx is None or code_idx is None:
            return jsonify({'error': f'无法识别列名，当前表头：{headers[:10]}'}), 400

        # 收集有效数据，跳过无效行
        batch = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= max(name_idx, code_idx):
                continue
            name = str(row[name_idx]).strip() if row[name_idx] else ''
            code = str(row[code_idx]).strip() if row[code_idx] else ''
            if code.isdigit():
                continue
            if name and code:
                batch.append((name, code))

        # 批量写入（500条/批），避免逐行网络往返
        BATCH_SIZE = 500
        total = 0
        for i in range(0, len(batch), BATCH_SIZE):
            chunk = batch[i:i + BATCH_SIZE]
            execute_values(cur,
                "INSERT INTO customers (name, code, level, last_followup, next_followup) VALUES %s ON CONFLICT (name) DO UPDATE SET code = EXCLUDED.code, updated_at = NOW()",
                [(n, c, 'D', None, None) for n, c in chunk],
                page_size=BATCH_SIZE
            )
            conn.commit()
            total += len(chunk)

        return jsonify({'success': True, 'message': f'已更新{total}个客户的编码'})
    finally:
        cur.close()
        conn.close()


@app.route('/api/customers/export-pending-codes')
def export_pending_codes():
    """导出待编码客户"""
    import csv
    import io
    conn = get_db()
    cur = get_cursor(conn)
    try:
        cur.execute("SELECT name, code FROM customers WHERE code IS NULL OR code = '' ORDER BY name")
        rows = cur.fetchall()
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['客户名称', '编码'])
        for row in rows:
            writer.writerow([row['name'], ''])
        output.seek(0)
        from flask import Response
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': 'attachment; filename=pending_codes.csv'}
        )
    finally:
        cur.close()
        conn.close()


# ── 启动：gunicorn 不走 __main__，需在模块加载时初始化数据库 ──
if DATABASE_URL:
    try:
        init_db()
        print(">>> DB initialized OK (module load)", flush=True)
    except Exception as e:
        print(f">>> DB INIT FAILED: {e}", flush=True)
        traceback.print_exc()

if __name__ == '__main__':
    print(">>> Starting sales dashboard...", flush=True)
    print(f">>> DATABASE_URL: {'***configured***' if DATABASE_URL else 'MISSING!'}", flush=True)
    try:
        init_db()
        print(">>> DB initialized OK", flush=True)
    except Exception as e:
        print(f">>> DB INIT FAILED: {e}", flush=True)
        traceback.print_exc()
    port = int(os.environ.get('PORT', 5000))
    print(f">>> Listening on 0.0.0.0:{port}", flush=True)
    app.run(host='0.0.0.0', port=port, debug=False)